import os
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO

EXCEL_FILE = "mo_oa.xlsx"  # Путь до excel файла с изображениями (если он в папке со скриптом то просто название файла)
IMG_OUTPUT_FOLDER = "img"  # Путь до папки, куда будут сохраняться изображения (если в папке со скриптом то просто название папки). При сохранении если файл уже будет в папке то перезапишется
SHEET = "BAZA"  # Имя листа с которого нужно забирать изображения
PHOTO_COLUMN = "Photo"  # Название столбца с изображениями
MDC_COLUMN = "MDC"  # Название столбца с артикулами, которые будут использоваться для названия изображений
HEADER_ROW = 5  # Номер строки с названиями столбцов (отсчет от 1)


def get_images(sheet: Worksheet, photo_col_idx: int, mdc_col_idx: int) -> None:
    # Получаем букву ячейки с артикулом
    mdc_column_letter = get_column_letter(mdc_col_idx)

    # Пробегаем по всем изображениями в файле, проверяем в каком столбце и строке находится изображение
    for image in sheet._images:
        row = image.anchor._from.row + 1
        col = image.anchor._from.col + 1

        # Если изображение в нужной нам колонке, продолжаем обработку (в файле могут быть изображения и в других колонках)
        if col == photo_col_idx:
            
            # Вытаскиваем название артикула, сразу приводим ко всем большим буквам
            mdc_value = sheet[f"{mdc_column_letter}{row}"].value.upper()

            # Если артикул пустой, то пропускаем, в консоль кидаем предупреждение
            if not mdc_value:
                print(
                    f"Ячейка {mdc_column_letter}{row} для артикула пустая, файл не сохранен!"
                )
                continue

            # Если артикул не пустой, получаем байтовое представление изображения
            image_data = image._data()
            image_bytes = BytesIO(image_data)

            # Кидаем объект в pillow
            pil_image = Image.open(image_bytes)

            # Если изображение в RGBA, переводим в RGB (необходимо для конвертации в JPG)
            if pil_image.mode == "RGBA":
                pil_image = pil_image.convert("RGB")

            # Сохраняем изображение в формате .jpg (т.к. у нас все картинки до этого также были в jpg)
            pil_image.save(f"./{IMG_OUTPUT_FOLDER}/{mdc_value}.jpg", "JPEG")


def main() -> None:
    try:
        os.makedirs(
            IMG_OUTPUT_FOLDER, exist_ok=True
        )  # Создаём папку для изображений, если её нет
        workbook = load_workbook(EXCEL_FILE, data_only=True)  # Открываем Excel файл
        sheet = workbook[SHEET]  # Выбираем лист
        headers = {
            cell.value.strip(): idx + 1 for idx, cell in enumerate(sheet[HEADER_ROW])
        }  # Создаем словарь с названием столбцов и их индексами
        photo_col_idx = headers.get(PHOTO_COLUMN)  # Получаем индекс столбца с изображениями
        mdc_col_idx = headers.get(MDC_COLUMN)  # Получаем индекс столбца с артикулами

        # Если не будут найдены нужные столбцы райзим исключение
        if not photo_col_idx or not mdc_col_idx:
            raise ValueError("Не удалось найти указанные столбцы в файле")

        get_images(sheet, photo_col_idx, mdc_col_idx)

    except Exception as error:
        print("Произошла ошибка:", error)


if __name__ == "__main__":
    main()
